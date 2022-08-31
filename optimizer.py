import pandas as pd
import openpyxl
from pulp import *
import re

wb = openpyxl.Workbook()
ws = wb.active

players = pd.read_csv(r"/Users/jordan/Desktop/J-G/Seedcopy.csv", usecols=['Name','Position','Salary','ProjPts'])

availables = players.groupby(["Position", "Name", "ProjPts", "Salary"]).agg('count')
availables = availables.reset_index()

salaries = {}
points = {}



for pos in availables.Position.unique():
    available_pos = availables[availables.Position == pos]
    salary = list(available_pos[['Name','Salary']].set_index('Name').to_dict().values())[0]
    point = list(available_pos[['Name','ProjPts']].set_index('Name').to_dict().values())[0]

    salaries[pos] = salary
    points[pos] = point

# print(salaries)
# print(points)

# RB in flex
pos_num_available = {
    "QB": 1,
    "RB": 3,
    "WR": 3,
    "TE": 1,
    "DST":1
}

salary_cap = 50000


for lineup in range(1,151):
    _vars = {k: LpVariable.dict(k, v, cat='Binary') for k, v in points.items()}

    prob = LpProblem("Fantasy", LpMaximize)
    rewards = []
    costs = []
    position_constraints = []

    for k, v in _vars.items():
        costs += lpSum([salaries[k][i] * _vars[k][i] for i in v])
        rewards += lpSum([points[k][i] * _vars[k][i] for i in v])
        prob += lpSum([_vars[k][i] for i in v]) == pos_num_available[k]

    prob += lpSum(rewards)
    prob += lpSum(costs) <= salary_cap
    if not lineup == 1:
        prob += (lpSum(rewards) <= total_score-0.01)
    prob.solve()

    score= str(prob.objective)
    constraints = [str(const) for const in prob.constraints.values()]
    colnum = 1

    for v in prob.variables():
        score = score.replace(v.name, str(v.varValue))
        if v.varValue !=0:
            ws.cell(row=lineup, column=colnum).value = v.name
            colnum +=1
    total_score = eval(score)
    ws.cell(row=lineup, column=colnum).value = total_score
    print(lineup, total_score)

wb.save(r"playerlists/2.xlsx")


